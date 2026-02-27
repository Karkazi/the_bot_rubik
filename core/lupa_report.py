"""
Логирование заявок Лупа (поиск на сайте) в Excel для отчёта администраторам.
Формат как в the_bot_lupa: data/lupa_tickets_log.xlsx.
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

STATS_FILE = Path(__file__).resolve().parents[2] / "data" / "lupa_tickets_log.xlsx"
HEADERS = ["Дата", "Время", "Автор (ФИО)", "Подразделение", "Табельный номер", "Задача в Jira"]


def init_stats_file() -> None:
    """Создаёт Excel-файл с заголовками, если его нет."""
    if not HAS_OPENPYXL:
        return
    STATS_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not STATS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки Лупа"
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        wb.save(str(STATS_FILE))


def log_lupa_ticket(
    channel_id: str,
    user_id: int,
    issue_key: str,
    full_name: Optional[str] = None,
    subdivision: Optional[str] = None,
    employee_id: Optional[str] = None,
) -> None:
    """
    Пишет одну строку в lupa_tickets_log.xlsx после создания заявки Лупа.
    """
    if not HAS_OPENPYXL or not issue_key:
        return
    now = datetime.now()
    author = (full_name or "").strip() or f"{channel_id}:{user_id}"
    subdivision = (subdivision or "").strip()
    employee_id = (employee_id or "").strip()
    row_data = [
        now.strftime("%d.%m.%Y"),
        now.strftime("%H:%M"),
        author,
        subdivision,
        employee_id,
        issue_key.strip().upper(),
    ]
    try:
        init_stats_file()
        wb = load_workbook(str(STATS_FILE))
        ws = wb.active
        next_row = ws.max_row + 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
        wb.save(str(STATS_FILE))
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Ошибка записи в отчёт Лупа: %s", e)


def get_report_path() -> Optional[Path]:
    """Путь к файлу отчёта, если он существует."""
    if STATS_FILE.exists():
        return STATS_FILE
    return None
